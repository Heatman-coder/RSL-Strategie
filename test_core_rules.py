import final
from core import candidate_engine
from core import console_ui as console_ui_core
from core import ranking as ranking_core
from core import quality_gate as quality_core
from core import reporting_excel as reporting_excel_core
import pandas as pd
from pathlib import Path
from uuid import uuid4


def _mk_stock(
    symbol: str,
    rank: int,
    rsl: float,
    sector: str = "Tech",
    industry: str = "Software",
    source_etf: str = "IVV",
    trust: int = 3,
    flag_stale: str = "OK",
    flag_gap: str = "OK",
    flag_liquidity: str = "OK",
    kurs: float = 100.0,
    sma: float = 90.0,
):
    return final.StockData(
        original_ticker=symbol,
        yahoo_symbol=symbol,
        isin="",
        name=symbol,
        sector=sector,
        industry=industry,
        land="US",
        market_value=float(1000 - rank),
        kurs=kurs,
        sma=sma,
        rsl=rsl,
        trust_score=trust,
        rsl_rang=rank,
        source_etf=source_etf,
        flag_stale=flag_stale,
        flag_gap=flag_gap,
        flag_liquidity=flag_liquidity,
    )


def test_normalize_sector_name_mappings():
    assert final.normalize_sector_name("Financial Institutions") == "Financials"
    assert final.normalize_sector_name("Industrial") == "Industrials"
    assert final.normalize_sector_name("Utility") == "Utilities"
    assert final.normalize_sector_name("Agency") == "Other"


def test_evaluate_rank_boundaries():
    top = 0.25
    res_hold = ranking_core.evaluate_rank(rank=21, universe_size=100, top_percent_threshold=top)
    res_warn = ranking_core.evaluate_rank(rank=22, universe_size=100, top_percent_threshold=top)
    res_sell = ranking_core.evaluate_rank(rank=26, universe_size=100, top_percent_threshold=top)

    assert res_hold["hold_rank"] == 25
    assert res_hold["warn_rank"] == 22
    assert res_hold["status"] == "HOLD"
    assert res_warn["status"] == "WARN"
    assert res_sell["status"] == "SELL"


def test_sort_portfolio_items_by_rank_best_to_worst():
    lookup = {
        "AAA": _mk_stock("AAA", rank=1, rsl=1.4),
        "BBB": _mk_stock("BBB", rank=3, rsl=1.2),
    }
    items = [
        {"Yahoo_Symbol": "BBB"},
        {"Yahoo_Symbol": "ZZZ"},
        {"Yahoo_Symbol": "AAA"},
    ]
    out = ranking_core.sort_portfolio_items_by_rank(items, lookup)
    assert [i["Yahoo_Symbol"] for i in out] == ["AAA", "BBB", "ZZZ"]


def test_build_multiscope_status_map_contains_expected_trigger_text():
    stocks = []
    for idx in range(1, 13):
        sym = f"S{idx:02d}"
        stocks.append(_mk_stock(sym, rank=idx, rsl=2.0 - idx * 0.01))

    status_map = ranking_core.build_multiscope_status_map(
        stock_results=stocks,
        top_percent_threshold=0.25,
        etf_options={"IVV": {"name": "iShares Core S&P 500 ETF"}},
    )

    assert status_map["S01"]["overall_status"] == "HOLD"
    assert status_map["S04"]["overall_status"] == "SELL"
    trigger_text = status_map["S04"]["trigger_scope_text"]
    assert "Gesamtliste" in trigger_text
    assert "Sektor: Tech" in trigger_text
    assert "ETF: IVV (iShares Core S&P 500 ETF)" in trigger_text


def test_percent_bar_color_thresholds():
    green = ranking_core.format_percent_bar(10.0)
    yellow = ranking_core.format_percent_bar(20.0)
    red = ranking_core.format_percent_bar(26.0)

    assert "\x1b[92m" in green
    assert "\x1b[93m" in yellow
    assert "\x1b[91m" in red
    assert "!" in red


def test_quality_report_and_strict_failure_eval():
    stocks = [
        _mk_stock("AAA", rank=1, rsl=1.5, trust=1, flag_gap="WARN"),
        _mk_stock("BBB", rank=2, rsl=1.4, flag_stale="WARN"),
        _mk_stock("CCC", rank=3, rsl=1.3, kurs=100.0, sma=0.0),
    ]
    report = quality_core.build_quality_report(
        stock_results=stocks,
        universe_candidates=10,
        failed_records_count=4,
        young_records_count=0,
        dropped_critical_count=0,
        portfolio_symbols=["AAA", "DDD"],
    )

    assert report["metrics"]["analyzed_count"] == 3
    assert report["metrics"]["invalid_numeric_count"] == 1
    assert report["ratios"]["coverage_ratio"] == 0.3
    assert report["ratios"]["portfolio_coverage_ratio"] == 0.5

    strict_cfg = {
        "strict_min_analyzed_stocks": 5,
        "strict_min_coverage_ratio": 0.5,
        "strict_max_failed_ratio": 0.2,
        "strict_max_young_ratio": 1.0,
        "strict_max_critical_drop_ratio": 1.0,
        "strict_max_stale_warn_ratio": 1.0,
        "strict_max_gap_warn_ratio": 1.0,
        "strict_max_liquidity_warn_ratio": 1.0,
        "strict_max_low_trust_ratio": 0.2,
        "strict_min_portfolio_coverage_ratio": 1.0,
        "strict_max_invalid_numeric_count": 0,
        "strict_max_duplicate_symbols": 0,
    }
    failures = quality_core.evaluate_strict_quality_failures(report, strict_cfg)
    assert failures
    assert any("Coverage" in msg or "analy" in msg.lower() for msg in failures)


def test_multiscope_warn_if_only_sector_is_warn():
    target = _mk_stock("WARNSEC", rank=3, rsl=1.90, sector="Energy", source_etf="QQQ")
    stocks = [target]

    for idx in range(3):
        stocks.append(_mk_stock(f"EHI{idx}", rank=100 + idx, rsl=2.20 - idx * 0.01, sector="Energy", source_etf="QQQ"))
    for idx in range(16):
        stocks.append(_mk_stock(f"ELO{idx}", rank=200 + idx, rsl=1.60 - idx * 0.01, sector="Energy", source_etf="QQQ"))
    for idx in range(20):
        stocks.append(_mk_stock(f"O{idx:02d}", rank=300 + idx, rsl=1.80 - idx * 0.01, sector="Tech", source_etf="QQQ"))

    status_map = ranking_core.build_multiscope_status_map(
        stock_results=stocks,
        top_percent_threshold=0.25,
        etf_options={"QQQ": {"name": "Invesco QQQ Trust"}},
    )

    info = status_map["WARNSEC"]
    assert info["overall_status"] == "WARN"
    assert info["trigger_scope_code"] == "S"
    assert "Sektor: Energy" in info["trigger_scope_text"]


def test_multiscope_sell_overrides_global_warn_if_etf_is_sell():
    target = _mk_stock("SELLETF", rank=9, rsl=1.70, sector="Tech", source_etf="SMALL")
    stocks = [target]

    for idx in range(7):
        stocks.append(_mk_stock(f"SM{idx}", rank=400 + idx, rsl=2.00 - idx * 0.03, sector="Tech", source_etf="SMALL"))
    for idx in range(32):
        stocks.append(_mk_stock(f"LG{idx:02d}", rank=500 + idx, rsl=1.60 - idx * 0.01, sector="Tech", source_etf="LARGE"))

    status_map = ranking_core.build_multiscope_status_map(
        stock_results=stocks,
        top_percent_threshold=0.25,
        etf_options={
            "SMALL": {"name": "Small Test ETF"},
            "LARGE": {"name": "Large Test ETF"},
        },
    )

    info = status_map["SELLETF"]
    assert info["overall_status"] == "SELL"
    assert info["trigger_scope_code"] == "E"
    assert "ETF: SMALL (Small Test ETF)" in info["trigger_scope_text"]


def test_sanitize_heatmap_thresholds_keeps_warn_below_full():
    warn, full = final._sanitize_heatmap_thresholds(30, 25)
    assert full == 25.0
    assert warn < full
    warn2, full2 = final._sanitize_heatmap_thresholds(-5, 0)
    assert warn2 >= 0.0
    assert full2 >= 1.0


def test_candidate_engine_prefers_top_ranked_candidates():
    config = dict(final.CONFIG)
    config.update(
        {
            "industry_top_n": 3,
            "industry_score_min": 0.0,
            "industry_breadth_min": 0.0,
            "industry_min_size": 1,
            "candidate_use_cluster_filter": False,
            "candidate_use_momentum_score": False,
            "candidate_use_vol_adjust": False,
            "candidate_use_industry_neutral": False,
            "candidate_use_accel": False,
            "candidate_use_rsl_change_1w": False,
            "candidate_min_avg_volume_eur": 0.0,
            "candidate_min_mktcap_m_eur": 0.0,
            "candidate_min_trust_score": 0,
            "candidate_excluded_countries": [],
            "candidate_score_min": 0.0,
            "candidate_require_top_percent": False,
            "candidate_top_percent_threshold": 0.25,
            "candidate_block_new_buys_in_weak_regime": False,
            "candidate_max_stocks_per_industry": 0,
            "candidate_use_peer_spread": False,
            "candidate_peer_spread_weight": 0.0,
            "candidate_max_distance_52w_high_pct": 0.0,
        }
    )
    industry_summary = final.summary_core.build_industry_rsl_summary(
        [
            _mk_stock("A1", rank=1, rsl=1.50, sector="Tech", industry="Software"),
            _mk_stock("A2", rank=2, rsl=1.45, sector="Tech", industry="Software"),
            _mk_stock("A3", rank=3, rsl=1.40, sector="Tech", industry="Software"),
            _mk_stock("B1", rank=4, rsl=1.35, sector="Energy", industry="Oil"),
        ],
        config,
    )
    stocks = [
        _mk_stock("A1", rank=1, rsl=1.50, sector="Tech", industry="Software"),
        _mk_stock("A2", rank=2, rsl=1.45, sector="Tech", industry="Software"),
        _mk_stock("A3", rank=3, rsl=1.40, sector="Tech", industry="Software"),
        _mk_stock("B1", rank=4, rsl=1.35, sector="Energy", industry="Oil"),
        _mk_stock("LATE", rank=90, rsl=1.05, sector="Tech", industry="Software"),
    ]
    symbol_lookup = {s.yahoo_symbol: s for s in stocks}

    out = candidate_engine.suggest_portfolio_candidates(
        stock_results=stocks,
        industry_summary=industry_summary,
        cluster_summary=None,
        portfolio_symbols=set(),
        sell_list_symbols=set(),
        symbol_lookup=symbol_lookup,
        config=config,
        portfolio_size=3,
    )

    assert [s.yahoo_symbol for s in out] == ["A1", "A2", "B1"]


def test_candidate_engine_can_return_debug_details():
    config = dict(final.CONFIG)
    config.update(
        {
            "industry_top_n": 3,
            "industry_score_min": 0.0,
            "industry_breadth_min": 0.0,
            "industry_min_size": 1,
            "candidate_use_cluster_filter": False,
            "candidate_use_momentum_score": False,
            "candidate_use_vol_adjust": False,
            "candidate_use_industry_neutral": False,
            "candidate_use_accel": True,
            "candidate_accel_weight": 0.5,
            "candidate_use_rsl_change_1w": False,
            "candidate_min_avg_volume_eur": 0.0,
            "candidate_min_mktcap_m_eur": 0.0,
            "candidate_min_trust_score": 0,
            "candidate_excluded_countries": [],
            "candidate_score_min": 0.0,
            "candidate_require_top_percent": False,
            "candidate_block_new_buys_in_weak_regime": False,
            "candidate_max_stocks_per_industry": 0,
            "candidate_use_peer_spread": False,
            "candidate_peer_spread_weight": 0.0,
            "candidate_max_distance_52w_high_pct": 0.0,
        }
    )

    a1 = _mk_stock("A1", rank=1, rsl=1.50, sector="Tech", industry="Software")
    a1.mom_accel = 0.10
    a2 = _mk_stock("A2", rank=2, rsl=1.40, sector="Energy", industry="Oil")
    a2.mom_accel = 0.00

    industry_summary = final.summary_core.build_industry_rsl_summary([a1, a2], config)
    symbol_lookup = {s.yahoo_symbol: s for s in [a1, a2]}

    stocks, details = candidate_engine.suggest_portfolio_candidates(
        stock_results=[a1, a2],
        industry_summary=industry_summary,
        cluster_summary=None,
        portfolio_symbols=set(),
        sell_list_symbols=set(),
        symbol_lookup=symbol_lookup,
        config=config,
        portfolio_size=2,
        return_details=True,
    )

    assert [s.yahoo_symbol for s in stocks] == ["A1", "A2"]
    detail_map = {item[1].yahoo_symbol: item[2] for item in details}
    assert detail_map["A1"]["symbol"] == "A1"
    assert detail_map["A1"]["base_label"] == "RSL-1"
    assert detail_map["A1"]["accel_component"] == 0.05
    assert "selection_reason" in detail_map["A1"]


def test_candidate_engine_blocks_new_buys_in_weak_market_regime():
    config = dict(final.CONFIG)
    config.update(
        {
            "industry_top_n": 3,
            "industry_score_min": 0.0,
            "industry_breadth_min": 0.0,
            "industry_min_size": 1,
            "candidate_use_cluster_filter": False,
            "candidate_use_momentum_score": False,
            "candidate_use_vol_adjust": False,
            "candidate_use_industry_neutral": False,
            "candidate_use_accel": False,
            "candidate_use_rsl_change_1w": False,
            "candidate_min_avg_volume_eur": 0.0,
            "candidate_min_mktcap_m_eur": 0.0,
            "candidate_min_trust_score": 0,
            "candidate_excluded_countries": [],
            "candidate_score_min": 0.0,
            "candidate_require_top_percent": False,
            "candidate_block_new_buys_in_weak_regime": True,
            "candidate_use_peer_spread": False,
            "candidate_peer_spread_weight": 0.0,
            "candidate_max_distance_52w_high_pct": 0.0,
        }
    )
    stocks = [
        _mk_stock("A1", rank=1, rsl=1.50, sector="Tech", industry="Software"),
        _mk_stock("A2", rank=2, rsl=1.45, sector="Energy", industry="Oil"),
    ]
    industry_summary = final.summary_core.build_industry_rsl_summary(stocks, config)
    symbol_lookup = {s.yahoo_symbol: s for s in stocks}

    out = candidate_engine.suggest_portfolio_candidates(
        stock_results=stocks,
        industry_summary=industry_summary,
        cluster_summary=None,
        portfolio_symbols=set(),
        sell_list_symbols=set(),
        symbol_lookup=symbol_lookup,
        config=config,
        market_regime={"regime": "SCHWACH"},
        portfolio_size=2,
    )

    assert out == []


def test_final_candidate_wrapper_accepts_console_style_keywords():
    out = final.suggest_portfolio_candidates(
        stock_results=[],
        industry_summary=None,
        cluster_summary=None,
        portfolio_symbols=set(),
        sell_list_symbols=set(),
        symbol_lookup={},
        config={},
        return_details=True,
    )

    assert out == []


def test_final_candidate_wrapper_accepts_legacy_config_second_arg():
    out = final.suggest_portfolio_candidates(
        [],
        {},
        industry_summary=None,
        cluster_summary=None,
        portfolio_symbols=set(),
        sell_list_symbols=set(),
        symbol_lookup={},
    )

    assert out == []


def test_candidate_engine_industry_cap_can_force_diversification():
    config = dict(final.CONFIG)
    config.update(
        {
            "industry_top_n": 3,
            "industry_score_min": 0.0,
            "industry_breadth_min": 0.0,
            "industry_min_size": 1,
            "candidate_use_cluster_filter": False,
            "candidate_use_momentum_score": False,
            "candidate_use_vol_adjust": False,
            "candidate_use_industry_neutral": False,
            "candidate_use_accel": False,
            "candidate_use_rsl_change_1w": False,
            "candidate_min_avg_volume_eur": 0.0,
            "candidate_min_mktcap_m_eur": 0.0,
            "candidate_min_trust_score": 0,
            "candidate_excluded_countries": [],
            "candidate_score_min": 0.0,
            "candidate_require_top_percent": False,
            "candidate_max_stocks_per_industry": 1,
            "candidate_block_new_buys_in_weak_regime": False,
            "candidate_use_peer_spread": False,
            "candidate_peer_spread_weight": 0.0,
            "candidate_max_distance_52w_high_pct": 0.0,
        }
    )
    stocks = [
        _mk_stock("A1", rank=1, rsl=1.50, sector="Tech", industry="Software"),
        _mk_stock("A2", rank=2, rsl=1.45, sector="Tech", industry="Software"),
        _mk_stock("B1", rank=3, rsl=1.40, sector="Energy", industry="Oil"),
    ]
    industry_summary = final.summary_core.build_industry_rsl_summary(stocks, config)
    symbol_lookup = {s.yahoo_symbol: s for s in stocks}

    out = candidate_engine.suggest_portfolio_candidates(
        stock_results=stocks,
        industry_summary=industry_summary,
        cluster_summary=None,
        portfolio_symbols=set(),
        sell_list_symbols=set(),
        symbol_lookup=symbol_lookup,
        config=config,
        portfolio_size=2,
    )

    assert [s.yahoo_symbol for s in out] == ["A1", "B1"]


def test_main_export_contains_atr_sell_limit():
    stock = _mk_stock("AAA", rank=1, rsl=1.5)
    stock.atr = 4.0
    stock.atr_limit = 96.0
    stock.atr_sell_limit = 100.6
    stock.peer_spread = 0.12
    stock.distance_52w_high_pct = 4.5

    df = console_ui_core._build_main_export_dataframe(
        stock_results=[stock],
        multiscope_status_map={},
        candidate_symbols=set(),
        watchlist_set=set(),
        build_yahoo_quote_url=lambda symbol: f"https://example.com/{symbol}",
    )

    assert "ATR" not in df.columns
    assert "ATR Buy" in df.columns
    assert "ATR Sell" in df.columns
    assert "Peer Spread" in df.columns
    assert "Abst. 52W-Hoch %" in df.columns
    assert "Link" not in df.columns
    assert float(df.loc[0, "ATR Buy"]) == 96.0
    assert float(df.loc[0, "ATR Sell"]) == 100.6
    assert float(df.loc[0, "Peer Spread"]) == 0.12
    assert float(df.loc[0, "Abst. 52W-Hoch %"]) == 4.5
    assert df.columns.tolist()[:11] == [
        "RSL",
        "Tr",
        "Ticker",
        "ISIN",
        "Name",
        "St",
        "Lk",
        "Sektor",
        "Branche",
        "Land",
        "Kurs",
    ]
    assert df.loc[0, "Lk"] == "D"
    assert "RSL-Rang" in df.columns


def test_resolve_market_value_from_sources_falls_back_to_info_marketcap():
    row = {"Market Value": 0}
    info = {"marketCap": 123_000_000}
    expected = 123_000_000 * final.CURRENCY_RATES["DEFAULT"]
    assert final._resolve_market_value_from_sources(row, info) == expected


def test_parse_etf_selection_input_accepts_symbols():
    etf_options = {
        "IVV": {"name": "iShares Core S&P 500 ETF"},
        "SOXX": {"name": "iShares Semiconductor ETF"},
    }
    assert final._parse_etf_selection_input("IVV, SOXX", etf_options) == ["IVV", "SOXX"]


def test_parse_etf_selection_input_accepts_indices_and_all():
    etf_options = {
        "IVV": {"name": "iShares Core S&P 500 ETF"},
        "SOXX": {"name": "iShares Semiconductor ETF"},
        "IWM": {"name": "iShares Russell 2000 ETF"},
    }
    assert final._parse_etf_selection_input("2,1", etf_options) == ["SOXX", "IVV"]
    assert final._parse_etf_selection_input("all", etf_options) == ["IVV", "SOXX", "IWM"]


def test_refresh_market_caps_for_relevant_exchange_stocks_uses_cached_marketcap():
    stock = _mk_stock("AAA.DE", rank=1, rsl=1.5, source_etf="IVV")
    stock.listing_source = "XETRA"
    stock.market_value = 0.0
    stock.market_cap = 0.0
    stock.avg_volume_eur = 500_000.0
    stock.trust_score = 3

    class _Mgr:
        def __init__(self):
            self.info_cache = {"AAA.DE": {"marketCap": 456_000_000}}
        def get_cached_info(self, ticker):
            return self.info_cache.get(ticker)
        def fetch_and_cache_info(self, ticker):
            return self.info_cache.get(ticker, {})
        def save_info_cache(self):
            pass

    final.refresh_market_caps_for_relevant_exchange_stocks([stock], _Mgr())
    assert stock.market_cap == 456_000_000
    assert stock.market_value == 456_000_000


def test_synchronize_portfolio_symbols_updates_stale_yahoo_symbols():
    tmp_dir = Path("tests") / "_tmp_portfolio_sync"
    tmp_dir.mkdir(exist_ok=True)
    portfolio_path = tmp_dir / f"{uuid4().hex}_portfolio.json"
    portfolio_path.write_text(
        '[{"Original_Ticker":"MU","Yahoo_Symbol":"MTE.DE","Name":"MICRON TECHNOLOGY INC"},{"Yahoo_Symbol":"N7T.F","Name":"NUTRIEN LTD"}]',
        encoding="utf-8",
    )
    mgr = final.PortfolioManager(str(portfolio_path))
    stocks = [
        _mk_stock("MU", rank=1, rsl=1.3, source_etf="IVV"),
        _mk_stock("NTR", rank=2, rsl=1.2, source_etf="ACWI"),
    ]
    stocks[0].name = "MICRON TECHNOLOGY INC"
    stocks[0].yahoo_symbol = "MU"
    stocks[1].name = "NUTRIEN LTD"
    stocks[1].yahoo_symbol = "NTR.TO"

    changed = final.synchronize_portfolio_symbols_with_stock_results(mgr, stocks)
    assert changed == 2
    assert [item["Yahoo_Symbol"] for item in mgr.current_portfolio] == ["MU", "NTR.TO"]


def test_build_history_symbol_overrides_prefers_native_or_xetra_over_frankfurt():
    raw_df = pd.DataFrame(
        [
            {"Ticker": "SHM.F", "Name": "Shimano Inc.", "ISIN": "", "Land": "Japan", "Source_ETF": "", "Listing_Source": "FRA"},
            {"Ticker": "7309.T", "Name": "Shimano Inc.", "ISIN": "", "Land": "Japan", "Source_ETF": "ACWI", "Listing_Source": ""},
            {"Ticker": "SHM.DE", "Name": "Shimano Inc.", "ISIN": "", "Land": "Japan", "Source_ETF": "", "Listing_Source": "XETRA"},
        ]
    )
    current_df = pd.DataFrame(
        [
            {"Ticker": "SHM.F", "Name": "Shimano Inc.", "ISIN": "", "Land": "Japan", "Source_ETF": "ACWI", "Listing_Source": "FRA"},
            {"Ticker": "SHM.DE", "Name": "Shimano Inc.", "ISIN": "", "Land": "Japan", "Source_ETF": "", "Listing_Source": "XETRA"},
        ]
    )
    overrides = final.build_history_symbol_overrides(raw_df, current_df)
    assert overrides["SHM.F"] == "7309.T"
    assert overrides["SHM.DE"] == "7309.T"


def test_build_history_symbol_overrides_matches_common_cross_listing_name_variants():
    raw_df = pd.DataFrame(
        [
            {"Ticker": "YEC0.F", "Name": "YASKAWA Electric Corporation", "ISIN": "US9850871057", "Land": "Japan", "Source_ETF": "", "Listing_Source": "FRA"},
            {"Ticker": "6506", "Name": "YASKAWA ELECTRIC CORP", "ISIN": "", "Land": "Japan", "Source_ETF": "IXUS", "Listing_Source": ""},
            {"Ticker": "J9R.F", "Name": "The Japan Steel Works, Ltd.", "ISIN": "JP3721400004", "Land": "Japan", "Source_ETF": "", "Listing_Source": "FRA"},
            {"Ticker": "5631", "Name": "JAPAN STEEL WORKS LTD", "ISIN": "", "Land": "Japan", "Source_ETF": "IEFA", "Listing_Source": ""},
            {"Ticker": "RL2.F", "Name": "Sumitomo Realty & Development Co., Ltd.", "ISIN": "JP3409000001", "Land": "Japan", "Source_ETF": "", "Listing_Source": "FRA"},
            {"Ticker": "8830", "Name": "SUMITOMO REALTY & DEVELOPMENT LTD", "ISIN": "", "Land": "Japan", "Source_ETF": "IEFA", "Listing_Source": ""},
            {"Ticker": "CJ1.F", "Name": "BlueNord ASA", "ISIN": "NO0010379266", "Land": "Norway", "Source_ETF": "", "Listing_Source": "FRA"},
            {"Ticker": "BNOR", "Name": "BLUENORD", "ISIN": "", "Land": "Norway", "Source_ETF": "IEUR", "Listing_Source": ""},
        ]
    )
    current_df = pd.DataFrame(
        [
            {"Ticker": "YEC0.F", "Name": "YASKAWA Electric Corporation", "ISIN": "US9850871057", "Land": "Japan", "Source_ETF": "", "Listing_Source": "FRA"},
            {"Ticker": "J9R.F", "Name": "The Japan Steel Works, Ltd.", "ISIN": "JP3721400004", "Land": "Japan", "Source_ETF": "", "Listing_Source": "FRA"},
            {"Ticker": "RL2.F", "Name": "Sumitomo Realty & Development Co., Ltd.", "ISIN": "JP3409000001", "Land": "Japan", "Source_ETF": "", "Listing_Source": "FRA"},
            {"Ticker": "CJ1.F", "Name": "BlueNord ASA", "ISIN": "NO0010379266", "Land": "Norway", "Source_ETF": "", "Listing_Source": "FRA"},
        ]
    )

    overrides = final.build_history_symbol_overrides(raw_df, current_df)

    assert overrides["YEC0.F"] == "6506.T"
    assert overrides["J9R.F"] == "5631.T"
    assert overrides["RL2.F"] == "8830.T"
    assert overrides["CJ1.F"] == "BNOR.OL"


def test_normalize_name_for_dedup_key_handles_foreign_company_suffixes():
    assert final._normalize_name_for_dedup_key("YASKAWA Electric Corporation") == "yaskawaelectric"
    assert final._normalize_name_for_dedup_key("YASKAWA ELECTRIC CORP") == "yaskawaelectric"
    assert final._normalize_name_for_dedup_key("The Japan Steel Works, Ltd.") == "japansteelworks"
    assert final._normalize_name_for_dedup_key("JAPAN STEEL WORKS LTD") == "japansteelworks"
    assert final._normalize_name_for_dedup_key("Sumitomo Realty & Development Co., Ltd.") == "sumitomorealtyanddevelopment"
    assert final._normalize_name_for_dedup_key("SUMITOMO REALTY & DEVELOPMENT LTD") == "sumitomorealtyanddevelopment"
    assert final._normalize_name_for_dedup_key("BlueNord ASA") == "bluenord"
    assert final._normalize_name_for_dedup_key("BLUENORD") == "bluenord"


def test_get_rsl_integrity_drop_reasons_rejects_active_secondary_history():
    reasons = final.get_rsl_integrity_drop_reasons(
        {
            "original_ticker": "J9R.F",
            "yahoo_symbol": "J9R.F",
            "land": "Japan",
            "rsl": 7.182982,
            "mom_6m": None,
            "mom_3m": None,
            "rsl_change_1w": -2.36,
            "trend_smoothness": 0.0,
            "flag_scale": "WARN",
        }
    )

    assert "secondary_history_not_allowed" in reasons
    assert "high_rsl_missing_mom6" in reasons
    assert "scale_flag_active" in reasons


def test_get_rsl_integrity_drop_reasons_rejects_high_rsl_without_6m_confirmation():
    reasons = final.get_rsl_integrity_drop_reasons(
        {
            "original_ticker": "VG",
            "yahoo_symbol": "VG",
            "land": "USA",
            "rsl": 1.514139,
            "mom_6m": 0.066820,
            "mom_3m": 1.082407,
            "rsl_change_1w": -0.237275,
            "trend_smoothness": 0.122835,
            "flag_scale": "OK",
        }
    )

    assert "high_rsl_without_6m_confirmation" in reasons
    assert "high_rsl_breakdown" in reasons


def test_build_home_market_rsl_audit_flags_unresolved_secondary_listing():
    stock = _mk_stock("J9R.F", rank=1, rsl=1.82, trust=1)
    stock.original_ticker = "J9R.F"
    stock.yahoo_symbol = "J9R.F"
    stock.name = "The Japan Steel Works, Ltd."
    stock.land = "Japan"
    stock.mom_6m = 0.03
    stock.mom_3m = 0.02
    stock.rsl_change_1w = 0.00
    stock.primary_liquidity_symbol = "5631.T"
    stock.primary_liquidity_basis = "ISIN"
    stock.listing_source = "FRA"
    stock.flag_scale = "OK"

    audit_df = final.build_home_market_rsl_audit([stock])

    assert len(audit_df) == 1
    row = audit_df.iloc[0]
    assert row["history_status"] == "SECONDARY_HISTORY_ACTIVE"
    assert bool(row["needs_review"]) is True
    assert bool(row["history_matches_home"]) is False
    assert "secondary_without_override" in row["review_reasons"]
    assert "secondary_history_active" in row["review_reasons"]
    assert "high_rsl_vs_weak_6m" in row["review_reasons"]


def test_build_home_market_rsl_audit_marks_home_override_as_ok():
    stock = _mk_stock("J9R.F", rank=42, rsl=1.04, trust=3)
    stock.original_ticker = "J9R.F"
    stock.yahoo_symbol = "5631.T"
    stock.name = "The Japan Steel Works, Ltd."
    stock.land = "Japan"
    stock.mom_6m = 0.08
    stock.mom_3m = 0.04
    stock.rsl_change_1w = 0.01
    stock.primary_liquidity_symbol = "5631.T"
    stock.primary_liquidity_basis = "ISIN"
    stock.listing_source = "FRA"
    stock.flag_scale = "OK"

    audit_df = final.build_home_market_rsl_audit([stock])

    assert len(audit_df) == 1
    row = audit_df.iloc[0]
    assert row["history_status"] == "OVERRIDDEN_TO_HOME"
    assert bool(row["needs_review"]) is False
    assert bool(row["history_matches_home"]) is True
    assert row["review_reasons"] == ""


def test_build_home_market_rsl_review_shortlist_focuses_on_top_ranks():
    top_stock = _mk_stock("J9R.F", rank=42, rsl=1.55, trust=1)
    top_stock.original_ticker = "J9R.F"
    top_stock.yahoo_symbol = "J9R.F"
    top_stock.name = "The Japan Steel Works, Ltd."
    top_stock.land = "Japan"
    top_stock.mom_6m = 0.02
    top_stock.primary_liquidity_symbol = "5631.T"
    top_stock.primary_liquidity_basis = "ISIN"

    lower_stock = _mk_stock("X8WA.F", rank=420, rsl=1.25, trust=1)
    lower_stock.original_ticker = "X8WA.F"
    lower_stock.yahoo_symbol = "X8WA.F"
    lower_stock.name = "Example Foreign Listing"
    lower_stock.land = "Japan"
    lower_stock.mom_6m = 0.01
    lower_stock.primary_liquidity_symbol = "1234.T"
    lower_stock.primary_liquidity_basis = "ISIN"

    audit_df = final.build_home_market_rsl_audit([top_stock, lower_stock])
    shortlist_df = final.build_home_market_rsl_review_shortlist(audit_df, top_rank=100)

    assert shortlist_df["original_ticker"].tolist() == ["J9R.F"]


def test_calculate_flags_marks_price_scale_break_as_critical():
    mgr = final.MarketDataManager("tests/_tmp_hist_cache.json", "tests/_tmp_info_cache.json")
    close = ([0.50] * 140) + ([20.0] * 60)
    hist = pd.DataFrame(
        {
            "Close": close,
            "High": [c * 1.01 for c in close],
            "Low": [c * 0.99 for c in close],
            "Volume": [1000] * len(close),
        }
    )
    flags = mgr._calculate_flags(hist, curr_price=20.0, sma=1.0)
    assert flags["flag_gap"] == "WARN"
    assert flags["flag_scale"] == "CRITICAL"
    assert flags["trust_score"] == 0


def test_get_history_batch_marks_cache_only_batches_as_no_network():
    mgr = final.MarketDataManager(
        f"tests/_tmp_hist_cache_{uuid4().hex}.json",
        f"tests/_tmp_info_cache_{uuid4().hex}.json",
    )
    version_str = mgr._get_cache_version_string()
    cache_key = f"AAA_{version_str}"
    mgr.cache[cache_key] = {
        "curr": 100.0,
        "sma": 90.0,
        "vol_eur": 1234.0,
        "flags": {"flag_gap": "OK", "trust_score": 3, "trend_sma50": "OK"},
    }

    result = mgr.get_history_batch(["AAA"])

    assert "AAA" in result
    assert mgr.last_history_batch_used_network is False


def test_get_history_batch_marks_missing_batches_as_network_use():
    mgr = final.MarketDataManager(
        f"tests/_tmp_hist_cache_{uuid4().hex}.json",
        f"tests/_tmp_info_cache_{uuid4().hex}.json",
    )
    original_download = final.yf.download
    today = pd.Timestamp.today().normalize()
    dates = pd.bdate_range(end=today, periods=220)
    fake_hist = pd.DataFrame(
        {
            "Open": [100.0] * len(dates),
            "High": [101.0] * len(dates),
            "Low": [99.0] * len(dates),
            "Close": [100.0] * len(dates),
            "Volume": [1000] * len(dates),
        },
        index=dates,
    )
    try:
        final.yf.download = lambda *args, **kwargs: fake_hist
        result = mgr.get_history_batch(["AAA"])
    finally:
        final.yf.download = original_download

    assert "AAA" in result
    assert mgr.last_history_batch_used_network is True


def test_main_excel_hides_expert_columns():
    stock = _mk_stock("AAA", rank=1, rsl=1.5)
    stock.mom_12m = 0.20
    stock.mom_6m = 0.10
    stock.mom_3m = 0.05
    stock.mom_score = 0.15
    stock.mom_vol = 0.30
    stock.mom_score_adj = 0.50
    stock.mom_accel = 0.02
    stock.mom_cluster = "2222"
    stock.trend_quality = "STAB"
    stock.trend_sma50 = 98.0
    stock.twss_score = 5.0
    stock.twss_raw_pct = 0.08
    stock.twss_date = "2026-03-20"
    stock.twss_days_ago = 1
    stock.atr_limit = 96.0
    stock.atr_sell_limit = 100.6

    df = console_ui_core._build_main_export_dataframe(
        stock_results=[stock],
        multiscope_status_map={},
        candidate_symbols=set(),
        watchlist_set=set(),
        build_yahoo_quote_url=lambda symbol: f"https://example.com/{symbol}",
    )

    report_dir = Path("tests") / "_tmp_reports"
    report_dir.mkdir(exist_ok=True)
    report_path = report_dir / f"{uuid4().hex}_main_hidden.xlsx"

    class _Logger:
        def info(self, *args, **kwargs):
            pass
        def warning(self, *args, **kwargs):
            pass
        def error(self, *args, **kwargs):
            pass

    ok = reporting_excel_core.save_excel_report_safely({"main": df}, str(report_path), _Logger())
    assert ok is True

    from openpyxl import load_workbook

    wb = load_workbook(report_path)
    ws = wb["main"]
    header_to_letter = {cell.value: cell.column_letter for cell in ws[1]}

    assert ws.freeze_panes == "G2"
    ticker_col = header_to_letter["Ticker"]
    assert ws[f"{ticker_col}2"].value == '=HYPERLINK("https://finance.yahoo.com/quote/AAA/?p=AAA","AAA")'
    assert ws[f"{header_to_letter['Lk']}2"].value == "D"
    assert ws.column_dimensions[header_to_letter["RSL-Rang"]].hidden is True
    assert ws.column_dimensions[header_to_letter["ETFs/Boerse"]].hidden is True
    assert ws.column_dimensions[header_to_letter["Mom 6M"]].hidden is True
    assert ws.column_dimensions[header_to_letter["Mom Score adj"]].hidden is True
    assert ws.column_dimensions[header_to_letter["Mom Cluster"]].hidden is True
    assert ws.column_dimensions[header_to_letter["RSL"]].hidden is False
    assert ws.column_dimensions[header_to_letter["ATR Sell"]].hidden is False
    assert ws.column_dimensions[header_to_letter["RSL"]].width <= 7.5
    assert ws.column_dimensions[header_to_letter["Tr"]].width <= 5.5
    assert ws.column_dimensions[header_to_letter["Lk"]].width <= 4.5


def test_main_excel_encodes_dotted_ticker_hyperlink():
    stock = _mk_stock("047040.KS", rank=1, rsl=1.5)

    df = console_ui_core._build_main_export_dataframe(
        stock_results=[stock],
        multiscope_status_map={},
        candidate_symbols=set(),
        watchlist_set=set(),
        build_yahoo_quote_url=lambda symbol: f"https://example.com/{symbol}",
    )

    report_dir = Path("tests") / "_tmp_reports"
    report_dir.mkdir(exist_ok=True)
    report_path = report_dir / f"{uuid4().hex}_main_dotted_link.xlsx"

    class _Logger:
        def info(self, *args, **kwargs):
            pass
        def warning(self, *args, **kwargs):
            pass
        def error(self, *args, **kwargs):
            pass

    ok = reporting_excel_core.save_excel_report_safely({"main": df}, str(report_path), _Logger())
    assert ok is True

    from openpyxl import load_workbook

    wb = load_workbook(report_path)
    ws = wb["main"]
    header_to_letter = {cell.value: cell.column_letter for cell in ws[1]}
    ticker_col = header_to_letter["Ticker"]

    assert ws[f"{header_to_letter['Lk']}2"].value == "S"
    assert (
        ws[f"{ticker_col}2"].value
        == '=HYPERLINK("https://www.google.com/search?q=site%3Afinance.yahoo.com%2Fquote%20%22047040.KS%22","047040.KS")'
    )


def test_build_raw_export_dataframe_contains_all_core_fields_and_context():
    stock = _mk_stock("AAA", rank=3, rsl=1.4, sector="Tech", industry="Software", source_etf="IVV,SOXX")
    stock.listing_source = "XETRA"
    stock.market_cap = 123456789.0
    stock.avg_volume_eur = 9876543.21
    stock.primary_liquidity_eur = 22222222.0
    stock.primary_liquidity_symbol = "AXTI"
    stock.primary_liquidity_basis = "NAME"
    stock.mom_12m = 0.25
    stock.mom_6m = 0.10
    stock.mom_3m = 0.04
    stock.mom_score = 0.18
    stock.mom_score_adj = 1.23
    stock.peer_spread = 0.11
    stock.distance_52w_high_pct = 2.5
    stock.flag_scale = "WARN"
    stock.scale_reason = "Recent/SMA ratio auffaellig (7.10)"
    stock.in_depot = "JA"

    df = console_ui_core._build_raw_export_dataframe(
        stock_results=[stock],
        multiscope_status_map={
            "AAA": {
                "overall_status": "HOLD",
                "primary_reason": "Okay",
                "reason_sell": "",
                "reason_warn": "",
                "trigger_scope_code": "-",
                "trigger_scope_text": "-",
                "pct_global": 0.04,
                "pct_sector": 0.08,
                "pct_industry": 0.12,
            }
        },
        candidate_symbols={"AAA"},
        watchlist_set={"AAA"},
        candidate_details_map={
            "AAA": {
                "final_score": 1.2345,
                "selection_reason": "Sektor-Slot frei",
                "penalties": {"dd": 0.1, "vol": 0.05},
            }
        },
    )

    assert "yahoo_symbol" in df.columns
    assert "flag_scale" in df.columns
    assert "multiscope_pct_global" in df.columns
    assert "candidate_final_score" in df.columns
    assert "candidate_penalties" in df.columns
    assert df.loc[0, "yahoo_symbol"] == "AAA"
    assert df.loc[0, "status_marker"] == "D"
    assert df.loc[0, "is_candidate"] == "JA"
    assert df.loc[0, "is_watchlist"] == "JA"
    assert df.loc[0, "listing_source"] == "XETRA"
    assert df.loc[0, "primary_liquidity_eur"] == 22222222.0
    assert df.loc[0, "primary_liquidity_symbol"] == "AXTI"
    assert df.loc[0, "candidate_final_score"] == 1.2345
    assert '"dd": 0.1' in str(df.loc[0, "candidate_penalties"])


def test_excel_export_includes_raw_data_sheet():
    stock = _mk_stock("AAA", rank=1, rsl=1.5)
    raw_df = console_ui_core._build_raw_export_dataframe(
        stock_results=[stock],
        multiscope_status_map={},
        candidate_symbols=set(),
        watchlist_set=set(),
        candidate_details_map={},
    )

    report_dir = Path("tests") / "_tmp_reports"
    report_dir.mkdir(exist_ok=True)
    report_path = report_dir / f"{uuid4().hex}_raw_sheet.xlsx"

    class _Logger:
        def info(self, *args, **kwargs):
            pass
        def warning(self, *args, **kwargs):
            pass
        def error(self, *args, **kwargs):
            pass

    ok = reporting_excel_core.save_excel_report_safely({"raw_data": raw_df}, str(report_path), _Logger())
    assert ok is True

    from openpyxl import load_workbook

    wb = load_workbook(report_path)
    assert "raw_data" in wb.sheetnames
    ws = wb["raw_data"]
    header_to_letter = {cell.value: cell.column_letter for cell in ws[1]}

    assert ws.freeze_panes == "A2"
    assert "yahoo_symbol" in header_to_letter
    assert (
        ws[f"{header_to_letter['yahoo_symbol']}2"].value
        == '=HYPERLINK("https://finance.yahoo.com/quote/AAA/?p=AAA","AAA")'
    )


def test_apply_primary_liquidity_context_prefers_better_linked_listing():
    fra = _mk_stock("AHV.F", rank=5, rsl=1.2, sector="Tech", industry="Hardware")
    fra.name = "AXT Inc."
    fra.avg_volume_eur = 21.0
    fra.listing_source = "FRA"

    us = _mk_stock("AXTI", rank=6, rsl=1.3, sector="Tech", industry="Hardware")
    us.name = "AXT Inc."
    us.avg_volume_eur = 7_894_920.0
    us.source_etf = "IVV"

    final.apply_primary_liquidity_context([fra, us])

    assert fra.primary_liquidity_symbol == "AXTI"
    assert us.primary_liquidity_symbol == "AXTI"
    assert fra.primary_liquidity_eur == 7_894_920.0
    assert us.primary_liquidity_eur == 7_894_920.0


def test_candidate_engine_uses_primary_liquidity_not_listing_turnover():
    stock = _mk_stock("AHV.F", rank=1, rsl=1.5, sector="Tech", industry="Hardware")
    stock.avg_volume_eur = 21.0
    stock.primary_liquidity_eur = 7_894_920.0
    stock.trust_score = 3
    stock.mom_score = 0.5
    stock.mom_score_adj = 1.5

    industry_summary = pd.DataFrame([{
        "Branche": "Hardware",
        "Score": 0.5,
        "Aktien": 10,
        "Breadth Ratio": 0.6,
    }])
    cluster_summary = pd.DataFrame()
    config = {
        "candidate_min_trust_score": 3,
        "candidate_min_avg_volume_eur": 1_000_000.0,
        "candidate_use_momentum_score": True,
        "candidate_use_vol_adjust": True,
        "candidate_use_industry_neutral": False,
        "candidate_use_accel": False,
        "candidate_use_rsl_change_1w": False,
        "candidate_use_peer_spread": False,
        "candidate_top_percent_threshold": 0.25,
        "candidate_require_top_percent": False,
        "industry_top_n": 10,
        "cluster_enabled": False,
        "candidate_use_cluster_filter": False,
        "candidate_block_new_buys_in_weak_regime": False,
        "candidate_max_stocks_per_industry": 0,
    }

    result = candidate_engine.suggest_portfolio_candidates(
        stock_results=[stock],
        industry_summary=industry_summary,
        cluster_summary=cluster_summary,
        portfolio_symbols=set(),
        sell_list_symbols=set(),
        symbol_lookup={"AHV.F": stock},
        config=config,
        market_regime={"regime": "NORMAL"},
        return_details=False,
    )

    assert len(result) == 1
    assert result[0].yahoo_symbol == "AHV.F"
