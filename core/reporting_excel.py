import os
import re
from typing import Any, Dict
from urllib.parse import quote
import numpy as np
import pandas as pd


def build_sector_color_map(df: pd.DataFrame, sector_col: str = "Sector") -> Dict[str, str]:
    if sector_col not in df.columns or df.empty:
        return {}

    sectors = (
        df[sector_col]
        .fillna("Unbekannt")
        .astype(str)
        .str.strip()
        .replace("", "Unbekannt")
    )

    ranked_sectors = list(sectors.value_counts().index)

    def _norm_sector_key(name: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(name).lower())

    preset_by_norm = {
        "informationtechnology": "D6E9FF",
        "communicationservices": "F8D8E8",
        "communication": "F8D8E8",
        "healthcare": "D9F2D9",
        "energy": "FFE0B2",
        "financials": "FFF2B2",
        "industrials": "E6E6FA",
        "consumerdiscretionary": "FDE2E4",
        "consumerstaples": "E2F0CB",
        "materials": "FDECC8",
        "utilities": "DDEBFF",
        "realestate": "EADCF8",
        "other": "EFEFEF",
        "unbekannt": "EFEFEF",
    }

    fallback_palette = [
        "D9E8FB",
        "FCE1D3",
        "DDF0DA",
        "E8DDF3",
        "FFF3C9",
        "F7DCE8",
        "D8F0EE",
        "FBE8D0",
        "E5ECFA",
        "E9F6E3",
    ]

    color_map: Dict[str, str] = {}
    fallback_idx = 0
    for sector in ranked_sectors:
        norm = _norm_sector_key(sector)
        color = preset_by_norm.get(norm)
        if not color:
            color = fallback_palette[fallback_idx % len(fallback_palette)]
            fallback_idx += 1
        color_map[sector] = color
    return color_map


def _build_yahoo_finance_url(symbol: str) -> str:
    raw_symbol = str(symbol or "").strip()
    if not raw_symbol:
        return ""
    if raw_symbol.isalnum():
        path_symbol = quote(raw_symbol, safe=".-_")
        query_symbol = quote(raw_symbol, safe=".-_")
        return f"https://finance.yahoo.com/quote/{path_symbol}/?p={query_symbol}"

    search_query = quote(f'site:finance.yahoo.com/quote "{raw_symbol}"', safe="")
    return f"https://www.google.com/search?q={search_query}"


def save_excel_report_safely(sheets: Dict[str, pd.DataFrame], filename: str, logger: Any, threshold_rank: int = -1) -> bool:
    try:
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except (ImportError, ModuleNotFoundError):
        logger.error("Excel-Export nicht moeglich: 'openpyxl' ist nicht installiert.")
        return False

    # Define formats
    FORMAT_PERCENT = "0.00%"
    FORMAT_PERCENT_1 = "0.0%"
    FORMAT_FLOAT_2 = "#,##0.00"
    FORMAT_FLOAT_4 = "#,##0.0000"

    # Map columns to formats
    column_formats = {
        "main": {
            "RSL": FORMAT_FLOAT_4,
            "RSL 1W Diff": FORMAT_FLOAT_4,
            "Mom 12M": FORMAT_PERCENT,
            "Mom 6M": FORMAT_PERCENT,
            "Mom 3M": FORMAT_PERCENT,
            "Mom Score": FORMAT_FLOAT_4,
            "Mom Vol 3M": FORMAT_FLOAT_4,
            "Mom Score adj": FORMAT_FLOAT_4,
            "Mom Accel": FORMAT_PERCENT,
            "Kurs": FORMAT_FLOAT_2,
            "ATR Buy": FORMAT_FLOAT_2,
            "ATR Sell": FORMAT_FLOAT_2,
            "Umsatz 20T (Mio EUR)": FORMAT_FLOAT_2,
            "Market Cap (Mio EUR)": FORMAT_FLOAT_2,
            "Peer Spread": FORMAT_FLOAT_4,
            "Abst. 52W-Hoch %": FORMAT_PERCENT_1,
            "Trend-Exzess": FORMAT_FLOAT_2,
            "Exzess-Max %": FORMAT_PERCENT_1,
        },
            "raw_data": {
                "market_value": "#,##0",
                "market_cap": "#,##0",
                "kurs": FORMAT_FLOAT_2,
                "sma": FORMAT_FLOAT_4,
                "rsl": FORMAT_FLOAT_4,
                "atr": FORMAT_FLOAT_2,
            "atr_limit": FORMAT_FLOAT_2,
            "atr_sell_limit": FORMAT_FLOAT_2,
            "avg_volume_eur": "#,##0.00",
            "primary_liquidity_eur": "#,##0.00",
            "rsl_change_1w": FORMAT_FLOAT_4,
            "mom_12m": FORMAT_PERCENT,
            "mom_6m": FORMAT_PERCENT,
            "mom_3m": FORMAT_PERCENT,
            "mom_score": FORMAT_FLOAT_4,
            "mom_vol": FORMAT_FLOAT_4,
            "mom_score_adj": FORMAT_FLOAT_4,
            "mom_accel": FORMAT_PERCENT,
            "industry_median_rsl": FORMAT_FLOAT_4,
            "peer_spread": FORMAT_FLOAT_4,
            "high_52w": FORMAT_FLOAT_2,
            "distance_52w_high_pct": FORMAT_PERCENT_1,
            "max_drawdown_6m": FORMAT_PERCENT,
            "trend_smoothness": FORMAT_FLOAT_4,
            "twss_score": FORMAT_FLOAT_2,
            "twss_raw_pct": FORMAT_PERCENT_1,
            "price_scale_ratio": FORMAT_FLOAT_4,
            "multiscope_pct_global": FORMAT_PERCENT_1,
            "multiscope_pct_sector": FORMAT_PERCENT_1,
            "multiscope_pct_industry": FORMAT_PERCENT_1,
            "re:^candidate_.*score.*": FORMAT_FLOAT_4,
            "re:^candidate_.*component$": FORMAT_FLOAT_4,
            "re:^candidate_.*pct$": FORMAT_PERCENT_1,
            "fallback_fraction": FORMAT_PERCENT_1,
        },
        "etf_summary": {
            "Durchschnitt RSL": FORMAT_FLOAT_4,
            "RSL am Grenzrang": FORMAT_FLOAT_4,
            "Top-% Schwelle": FORMAT_PERCENT,
            "Breadth Ratio": FORMAT_PERCENT,
            "Strong Breadth Ratio": FORMAT_PERCENT,
            "Leader Ratio": FORMAT_PERCENT,
            "Score": FORMAT_FLOAT_4,
            "re:^Score .*": FORMAT_FLOAT_4,
        },
        "sector_summary": {
            "Durchschnitt RSL": FORMAT_FLOAT_4,
            "RSL am Grenzrang": FORMAT_FLOAT_4,
            "Top-% Schwelle": FORMAT_PERCENT,
        },
        "industry_summary": {
            "Breadth Ratio": FORMAT_PERCENT,
            "Strong Breadth Ratio": FORMAT_PERCENT,
            "Leader Ratio": FORMAT_PERCENT,
            "Avg RSL": FORMAT_FLOAT_4,
            "Median RSL": FORMAT_FLOAT_4,
            "Top RSL": FORMAT_FLOAT_4,
            "Score": FORMAT_FLOAT_4,
            "re:^Score .*": FORMAT_FLOAT_4,  # Regex for 'Score (vor XW)' etc.
        },
        "cluster_summary": {
            "Avg_Mom12": FORMAT_PERCENT,
            "Avg_Mom6": FORMAT_PERCENT,
            "Avg_Mom3": FORMAT_PERCENT,
            "Avg_Accel": FORMAT_PERCENT,
            "Avg_RSL": FORMAT_FLOAT_4,
            "Cluster Anteil %": FORMAT_PERCENT_1,
            "Score": FORMAT_FLOAT_4,
        },
        "universe_audit": {
            "Status": Font(bold=True),
        },
        "integrity_issues": {
            "rsl": FORMAT_FLOAT_4,
            "fallback_fraction": FORMAT_PERCENT_1,
        },
    }

    while True:
        try:
            hidden_main_cols = {
                "RSL-Rang",
                "ETFs",
                "ETFs/Boerse",
                "Orig. Ticker",
                "MktCap-Rang",
                "RSL 1W Diff",
                "Mom Cluster",
                "Mom 6M",
                "Mom 3M",
                "Mom Score",
                "Mom Vol 3M",
                "Mom Score adj",
                "Mom Accel",
                "SMA50",
                "Trust-Details",
            }
            main_df = sheets.get("main")
            held_symbols: set = set()
            held_sectors: set = set()
            sector_color_map_main: Dict[str, str] = {}
            if isinstance(main_df, pd.DataFrame) and not main_df.empty:
                if "Sektor" in main_df.columns:
                    sector_color_map_main = build_sector_color_map(main_df, sector_col="Sektor")
                if "St" in main_df.columns:
                    held_mask = main_df["St"].astype(str).str.strip().str.upper() == "D"
                    held_df = main_df[held_mask]
                    if "Ticker" in held_df.columns:
                        held_symbols = set(
                            held_df["Ticker"]
                            .astype(str)
                            .str.strip()
                            .str.upper()
                            .replace("", np.nan)
                            .dropna()
                            .tolist()
                        )
                    if "Sektor" in held_df.columns:
                        held_sectors = set(
                            held_df["Sektor"]
                            .astype(str)
                            .str.strip()
                            .replace("", np.nan)
                            .dropna()
                            .tolist()
                        )

            hold_row_default_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
            etf_hold_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
            sector_hold_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
            candidate_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")  # Hellgruen fuer Kaufkandidaten
            industry_top_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
            buy_cut_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")  # Gruen fuer BUY_CUT Zeile
            threshold_fill = PatternFill(fill_type="solid", fgColor="FF9999")  # Rot fuer die Cutoff-Zeile
            watchlist_fill = PatternFill(fill_type="solid", fgColor="E0F7FF")
            role_font_map = {
                "Depot-Ticker (Yahoo)": Font(color="1F4E78", bold=False),
                "Kaufkandidaten (Yahoo)": Font(color="008000", bold=True),
                "Verkaufssignale (Yahoo)": Font(color="C00000", bold=True),
            }
            main_width_map = {
                "RSL": 7,
                "Tr": 5,
                "Ticker": 12,
                "Lk": 4,
                "Name": 26,
                "St": 4,
                "Sektor": 16,
                "Branche": 24,
                "Land": 8,
                "Kurs": 10,
                "ATR Buy": 10,
                "ATR Sell": 10,
                "Market Cap (Mio EUR)": 15,
                "Umsatz 20T (Mio EUR)": 15,
                "Mom 12M": 10,
                "Trust": 6,
                "Trend-Qual.": 10,
                "Neu?": 5,
                "Erfasst seit": 12,
            }
            raw_width_map = {
                "yahoo_symbol": 14,
                "original_ticker": 14,
                "isin": 16,
                "name": 30,
                "sector": 18,
                "industry": 28,
                "land": 12,
                "source_etf": 22,
                "listing_source": 12,
                "primary_liquidity_symbol": 16,
                "primary_liquidity_basis": 10,
                "multiscope_primary_reason": 40,
                "multiscope_trigger_scope_text": 28,
                "scale_reason": 34,
            }

            def _fill_row(ws, row_idx: int, max_col: int, fill_obj):
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill_obj

            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                for sheet_name, df in sheets.items():
                    excel_sheet_name = str(sheet_name)[:31]
                    df.to_excel(writer, sheet_name=excel_sheet_name, index=False)
                    ws = writer.sheets[excel_sheet_name]
                    ws.freeze_panes = "G2" if excel_sheet_name == "main" else "A2"
                    ws.auto_filter.ref = ws.dimensions
                    ws.row_dimensions[1].height = 22
                    ws.sheet_format.defaultRowHeight = 18

                    for col_idx, col_name in enumerate(df.columns, start=1):
                        header_len = len(str(col_name))
                        if len(df.index) > 0:
                            col_values = df.iloc[:, col_idx - 1].tolist()
                            data_len = max(
                                [
                                    len(str(v))
                                    for v in col_values
                                    if v is not None and not (isinstance(v, float) and np.isnan(v))
                                ]
                                + [0]
                            )
                        else:
                            data_len = 0

                        best_width = min(max(10, header_len, data_len) + 2, 80)
                        ws.column_dimensions[get_column_letter(col_idx)].width = best_width

                    if excel_sheet_name == "main":
                        for col_idx, col_name in enumerate(df.columns, start=1):
                            width = main_width_map.get(str(col_name))
                            if width is not None:
                                ws.column_dimensions[get_column_letter(col_idx)].width = width

                        for col_idx, col_name in enumerate(df.columns, start=1):
                            if col_name in hidden_main_cols:
                                ws.column_dimensions[get_column_letter(col_idx)].hidden = True

                        if "Sektor" in df.columns and len(df.index) > 0:
                            sector_col_idx = list(df.columns).index("Sektor") + 1
                            sector_color_map = sector_color_map_main or build_sector_color_map(df, sector_col="Sektor")
                            status_col_idx = list(df.columns).index("St") + 1 if "St" in df.columns else None
                            ysym_col_idx = (
                                list(df.columns).index("Ticker") + 1 if "Ticker" in df.columns else None
                            )
                            for row_idx in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row_idx, column=sector_col_idx)
                                sector_name = str(cell.value).strip() if cell.value is not None else ""
                                if not sector_name:
                                    sector_name = "Unbekannt"
                                color = sector_color_map.get(sector_name)
                                is_held = False

                                status_val = ""
                                if status_col_idx:
                                    status_val = str(ws.cell(row=row_idx, column=status_col_idx).value or "").strip().upper()

                                is_held = status_val.startswith("D")
                                is_candidate = "K" in status_val.split("/")
                                is_buy_cut = "BUY_CUT" in status_val
                                is_watchlist = status_val.startswith("W")

                                if not is_held and ysym_col_idx is not None:
                                    ysym = str(ws.cell(row=row_idx, column=ysym_col_idx).value or "").strip().upper()
                                    is_held = ysym in held_symbols

                                if is_buy_cut:
                                    _fill_row(ws, row_idx, ws.max_column, buy_cut_fill)
                                elif is_held:
                                    # Einheitliche Farbe fuer Depotwerte, unabhaengig vom Sektor
                                    _fill_row(ws, row_idx, ws.max_column, hold_row_default_fill)
                                elif is_candidate:
                                    _fill_row(ws, row_idx, ws.max_column, candidate_fill)
                                elif is_watchlist:
                                    _fill_row(ws, row_idx, ws.max_column, watchlist_fill)
                                elif color:
                                    cell.fill = PatternFill(fill_type="solid", fgColor=color)

                        if "Ticker" in df.columns and len(df.index) > 0:
                            info_col_idx = list(df.columns).index("Ticker") + 1
                            for row_idx in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row_idx, column=info_col_idx)
                                ticker = str(cell.value).strip() if cell.value is not None else ""
                                if ticker:
                                    url = _build_yahoo_finance_url(ticker)
                                    display = ticker.replace('"', '""')
                                    cell.value = f'=HYPERLINK("{url}","{display}")'
                                    cell.style = "Hyperlink"

                        # Markiere die Threshold-Zeile (Top-X% Grenze) am Ende, um vorherige Fills zu ueberschreiben
                        if threshold_rank > 0 and "RSL-Rang" in df.columns:
                            rank_col_idx = list(df.columns).index("RSL-Rang") + 1
                            for row_idx in range(2, ws.max_row + 1):
                                rank_val = ws.cell(row=row_idx, column=rank_col_idx).value
                                try:
                                    if rank_val is not None and int(rank_val) == threshold_rank:
                                        _fill_row(ws, row_idx, ws.max_column, threshold_fill)
                                except (ValueError, TypeError):
                                    pass

                    if excel_sheet_name == "raw_data":
                        for col_idx, col_name in enumerate(df.columns, start=1):
                            width = raw_width_map.get(str(col_name))
                            if width is not None:
                                ws.column_dimensions[get_column_letter(col_idx)].width = width

                        if "yahoo_symbol" in df.columns and len(df.index) > 0:
                            info_col_idx = list(df.columns).index("yahoo_symbol") + 1
                            for row_idx in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row_idx, column=info_col_idx)
                                ticker = str(cell.value).strip() if cell.value is not None else ""
                                if ticker:
                                    url = _build_yahoo_finance_url(ticker)
                                    display = ticker.replace('"', '""')
                                    cell.value = f'=HYPERLINK("{url}","{display}")'
                                    cell.style = "Hyperlink"

                    for col_idx, col_name in enumerate(df.columns, start=1):
                        font_obj = role_font_map.get(str(col_name))
                        if font_obj is None:
                            continue
                        for row_idx in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value not in (None, ""):
                                cell.font = font_obj

                    # Apply number formats
                    sheet_formats = column_formats.get(excel_sheet_name, {})
                    if sheet_formats and len(df.index) > 0:
                        for col_idx, col_name in enumerate(df.columns, start=1):
                            fmt = sheet_formats.get(col_name)
                            if not fmt:
                                for re_key, re_fmt in sheet_formats.items():
                                    if re_key.startswith("re:") and re.match(re_key[3:], col_name):
                                        fmt = re_fmt
                                        break
                            if fmt:
                                for row_idx in range(2, ws.max_row + 1):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if isinstance(cell.value, (int, float, np.number)):
                                        cell.number_format = fmt

                    if (
                        excel_sheet_name == "industry_summary"
                        and len(df.index) > 0
                        and "Top-Kandidat-Branche" in df.columns
                    ):
                        top_col_idx = list(df.columns).index("Top-Kandidat-Branche") + 1
                        ws.column_dimensions[get_column_letter(top_col_idx)].hidden = True
                        for row_idx in range(2, ws.max_row + 1):
                            flag_val = str(ws.cell(row=row_idx, column=top_col_idx).value or "").strip().upper()
                            if flag_val == "JA":
                                _fill_row(ws, row_idx, ws.max_column, industry_top_fill)

                    if (
                        excel_sheet_name == "sector_summary"
                        and held_sectors
                        and len(df.index) > 0
                        and "Sektor" in df.columns
                    ):
                        sector_col_idx = list(df.columns).index("Sektor") + 1
                        for row_idx in range(2, ws.max_row + 1):
                            sector_name = str(ws.cell(row=row_idx, column=sector_col_idx).value or "").strip()
                            if sector_name in held_sectors:
                                _fill_row(ws, row_idx, ws.max_column, sector_hold_fill)
            logger.info(f"Gespeichert: {filename}")
            return True
        except PermissionError:
            print(f"\nACHTUNG: Die Datei '{os.path.basename(filename)}' ist noch geoeffnet!")
            print("--> Bitte schliessen Sie die Datei in Excel/Editor.")
            user_in = input("--> Druecken Sie ENTER, um es erneut zu versuchen (oder 'x' zum Abbrechen): ")
            if user_in.strip().lower() == "x":
                logger.warning("Excel-Speichern durch Benutzer abgebrochen.")
                return False
        except ValueError as e:
            logger.error(f"Excel-Export fehlgeschlagen: {e}")
            return False
        except Exception as e:
            logger.error(f"Kritischer Fehler beim Speichern von {filename}: {e}")
            return False
